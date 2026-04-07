"""Analytics module for generating reports."""
from datetime import date
from typing import Optional
from pathlib import Path
import matplotlib.pyplot as plt
import matplotlib
from sqlalchemy.orm import Session

from ..database import crud
from ..database.models import Earning, Expense
from ..utils import (
    format_currency, format_date, get_date_range,
    get_ist_today
)

matplotlib.use('Agg')


async def generate_report(
    format: str,
    period: str,
    include_charts: bool = True,
    db_session: Session = None
) -> str:
    """Generate a report in PDF or Excel format."""
    if not db_session:
        raise ValueError("Database session required")
    
    start_date, end_date = get_date_range(period)

    earnings = await crud.EarningCRUD.get_by_date_range(
        db=db_session,
        start_date=start_date,
        end_date=end_date
    )

    expenses = await crud.ExpenseCRUD.get_by_date_range(
        db=db_session,
        start_date=start_date,
        end_date=end_date
    )
    
    if format == "pdf":
        return await generate_pdf_report(
            earnings=earnings,
            expenses=expenses,
            period=period,
            start_date=start_date,
            end_date=end_date,
            include_charts=include_charts
        )
    elif format == "excel":
        return await generate_excel_report(
            earnings=earnings,
            expenses=expenses,
            period=period,
            start_date=start_date,
            end_date=end_date
        )
    else:
        raise ValueError(f"Unsupported format: {format}")


async def generate_pdf_report(
    earnings,
    expenses,
    period: str,
    start_date: date,
    end_date: date,
    include_charts: bool = True
) -> str:
    """Generate PDF report."""
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    
    output_dir = Path("./reports")
    output_dir.mkdir(exist_ok=True)
    
    filename = f"report_{period.replace(' ', '_')}_{get_ist_today().strftime('%Y%m%d')}.pdf"
    filepath = output_dir / filename
    
    doc = SimpleDocTemplate(str(filepath), pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.darkblue,
        spaceAfter=30
    )
    
    story.append(Paragraph(f"Financial Report - {period.title()}", title_style))
    story.append(Spacer(1, 12))
    
    story.append(Paragraph(f"Period: {format_date(start_date)} to {format_date(end_date)}", styles['Normal']))
    story.append(Spacer(1, 12))
    
    total_earnings = sum(e.amount_earned for e in earnings)
    total_expenses = sum(e.amount for e in expenses)
    net_income = total_earnings - total_expenses
    
    summary_data = [
        ["Metric", "Amount"],
        ["Total Earnings", format_currency(total_earnings)],
        ["Total Expenses", format_currency(total_expenses)],
        ["Net Income", format_currency(net_income)],
        ["Number of Earnings", str(len(earnings))],
        ["Number of Expenses", str(len(expenses))]
    ]
    
    summary_table = Table(summary_data, colWidths=[2*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(summary_table)
    story.append(Spacer(1, 24))
    
    if earnings:
        story.append(Paragraph("Earnings Details", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        earnings_data = [["Brand", "Amount", "Date", "Type", "Status"]]
        for e in earnings[:50]:
            earnings_data.append([
                e.brand_name,
                format_currency(e.amount_earned),
                format_date(e.entry_date),
                e.payment_type.value,
                e.status.value
            ])
        
        earnings_table = Table(earnings_data, colWidths=[1.5*inch, 1*inch, 1*inch, 0.8*inch, 0.8*inch])
        earnings_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(earnings_table)
        story.append(Spacer(1, 24))
    
    if expenses:
        story.append(Paragraph("Expenses Details", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        expenses_data = [["Category", "Amount", "Date", "Description"]]
        for e in expenses[:50]:
            expenses_data.append([
                e.category.value,
                format_currency(e.amount),
                format_date(e.expense_date),
                e.description or ""
            ])
        
        expenses_table = Table(expenses_data, colWidths=[1.2*inch, 1*inch, 1*inch, 2*inch])
        expenses_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(expenses_table)
        story.append(Spacer(1, 24))
    
    if include_charts and earnings:
        chart_path = await generate_earnings_chart(earnings, period)
        if chart_path and Path(chart_path).exists():
            story.append(Paragraph("Earnings by Brand", styles['Heading2']))
            story.append(Spacer(1, 12))
            img = Image(chart_path, width=5*inch, height=3*inch)
            story.append(img)
            story.append(Spacer(1, 24))
    
    doc.build(story)
    
    return str(filepath)


async def generate_excel_report(
    earnings,
    expenses,
    period: str,
    start_date: date,
    end_date: date
) -> str:
    """Generate Excel report."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.chart import BarChart, Reference
    
    output_dir = Path("./reports")
    output_dir.mkdir(exist_ok=True)
    
    filename = f"report_{period.replace(' ', '_')}_{get_ist_today().strftime('%Y%m%d')}.xlsx"
    filepath = output_dir / filename
    
    wb = Workbook()
    
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    total_earnings = sum(e.amount_earned for e in earnings)
    total_expenses = sum(e.amount for e in expenses)
    net_income = total_earnings - total_expenses
    
    headers = ["Metric", "Value"]
    ws_summary.append(headers)
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for cell in ws_summary[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
    
    ws_summary.append(["Total Earnings", total_earnings])
    ws_summary.append(["Total Expenses", total_expenses])
    ws_summary.append(["Net Income", net_income])
    ws_summary.append(["Number of Earnings", len(earnings)])
    ws_summary.append(["Number of Expenses", len(expenses)])
    ws_summary.append(["Period Start", format_date(start_date)])
    ws_summary.append(["Period End", format_date(end_date)])
    
    for row in ws_summary.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='left')
    
    if earnings:
        ws_earnings = wb.create_sheet("Earnings")
        
        earning_headers = ["ID", "Brand", "Amount", "Payment Type", "Date", "Status", "Notes"]
        ws_earnings.append(earning_headers)
        
        for cell in ws_earnings[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")
        
        for e in earnings:
            ws_earnings.append([
                e.id,
                e.brand_name,
                e.amount_earned,
                e.payment_type.value,
                format_date(e.entry_date),
                e.status.value,
                e.notes or ""
            ])
        
        for column in ws_earnings.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (AttributeError, TypeError):
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_earnings.column_dimensions[column_letter].width = adjusted_width
    
    if expenses:
        ws_expenses = wb.create_sheet("Expenses")
        
        expense_headers = ["ID", "Category", "Amount", "Date", "Description"]
        ws_expenses.append(expense_headers)
        
        for cell in ws_expenses[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")
        
        for e in expenses:
            ws_expenses.append([
                e.id,
                e.category.value,
                e.amount,
                format_date(e.expense_date),
                e.description or ""
            ])
        
        for column in ws_expenses.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (AttributeError, TypeError):
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_expenses.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(str(filepath))
    
    return str(filepath)


async def generate_earnings_chart(earnings, period: str) -> Optional[str]:
    """Generate earnings by brand chart."""
    if not earnings:
        return None
    
    brand_totals = {}
    for e in earnings:
        if e.brand_name not in brand_totals:
            brand_totals[e.brand_name] = 0
        brand_totals[e.brand_name] += e.amount_earned
    
    sorted_brands = sorted(brand_totals.items(), key=lambda x: x[1], reverse=True)[:10]
    
    if not sorted_brands:
        return None
    
    brands = [item[0] for item in sorted_brands]
    amounts = [item[1] for item in sorted_brands]
    
    plt.figure(figsize=(10, 6))
    bars = plt.bar(brands, amounts, color='steelblue')
    
    plt.xlabel('Brand')
    plt.ylabel('Amount (₹)')
    plt.title(f'Earnings by Brand - {period.title()}')
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()
    
    for bar in bars:
        height = bar.get_height()
        plt.text(bar.get_x() + bar.get_width()/2., height,
                f'₹{int(height)}',
                ha='center', va='bottom')
    
    output_dir = Path("./reports")
    output_dir.mkdir(exist_ok=True)
    
    chart_filename = f"earnings_chart_{period.replace(' ', '_')}_{get_ist_today().strftime('%Y%m%d')}.png"
    chart_path = output_dir / chart_filename
    
    plt.savefig(chart_path, dpi=150, bbox_inches='tight')
    plt.close()
    
    return str(chart_path)
